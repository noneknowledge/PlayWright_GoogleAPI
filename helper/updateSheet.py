from openpyxl import load_workbook
import os
from datetime import datetime

def getUpdateSheet(fileName:str , historyContent:str = None, historyDate = None, sheetIndex:int = 0 ,row:int = 2,):
    if os.path.exists(f"./data/{fileName}.xlsx"):
        wb = load_workbook(f"./data/{fileName}.xlsx")
        ws = wb.worksheets[sheetIndex]
        now = datetime.now()
        ws[f'E{row}'] = "✅"
        ws[f'F{row}'] = now.strftime("%Y-%m-%d %H:%M:%S")
        ws[f'G{row}'] = historyContent
        ws[f'H{row}'] = historyDate
        wb.save(f"./data/{fileName}.xlsx")
        print("Done update")
    else:
        print("Not Exist")
    return 


